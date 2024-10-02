'''from time import sleep
import pandas as pd
import random 
print('^'*40)
print(f' Jogos loteria'.center(40))
print('^'*40)
print()
lista = ['\033[1;33mquina\033[m','\033[1;34mmega-sena\033[m','\033[1;35mlotofacil\033[m','\033[1;31mSair\033[m']
for k,v in enumerate(lista):
    print(f'{k+1} - {v}')
while True:
    resposta =int(input('Qual jogo deseja fazer:'))
    if resposta == 1:
        cont = 0
        print('\033[1;43mVoce escolheu Quina\033[m')
        quant = int(input('Quantos jogos deseja fazer : '))
        sleep(1)
        print('O jogo sorteado foi :')
        while cont < quant :   
            jogo= random.sample(range(1,81),5)
            for n in jogo:
                print(f'\033[1;32m{n} - \033[m',end='')       
            cont += 1
            print()
        print(f'O valor dos jogos fica \033[1;31mR${2.5*quant:.2f}\033[m')    
    if resposta == 2:
        cont = 0
        print('\033[1;44mVoce escolheu Mega - Sena\033[m')
        sleep(1)
        quant = int(input('Quantos jogos quer fazer : '))
        sleep(1)
        print('O jogo sorteado foi :')
        while cont < quant :
            jogo= random.sample(range(1,61),6)
            for n in jogo:
                print (f'\033[1;32m{n} - \033[m',end='')
            cont += 1
            print()
        print(f'O valor dos jogos fica \033[1;31mR${5*quant:.2f}\033[m')
    if resposta == 3:
        cont = 0
        print(f'\033[1;45m{"Voce escolheu lotofacil"}\033[m')
        quant = int(input('Quantos jogos deseja fazer :'))
        sleep(1)
        print('O jogo sorteado foi :')
        while cont < quant :    
            jogo = random.sample(range(1,26),15)
            for n in jogo:
                print (f'\033[1;32m{n} - \033[m',end='')
            cont += 1
            print()
        print(f'O valor dos jogos fica \033[1;31mR${3*quant:.2f}\033[m')    
    if resposta == 4 :
        break '''

'''from openpyxl import Workbook,load_workbook


planilha = Workbook()
planilha_ativa = planilha.active

planilha_ativa['A1'] = 'id'
planilha_ativa['B1'] = 'produto' 
planilha_ativa['C1'] = 'quantidade'
planilha_ativa['D1'] = 'Preço_bruto'
planilha_ativa['E1'] = 'preço_final'
planilha_ativa['F1'] = 'fornecedor'
planilha_ativa['G1'] = 'data_de_entrada'
planilha_ativa['H1'] = 'data_de_validade'
planilha_ativa['I1'] = 'categoria'
planilha_ativa['J1'] = 'codigo'
planilha_ativa['K1'] = 'Total de produtos'
planilha.save("controle_de_estoque_bombodega.xlsx")

wb=load_workbook("controle_de_estoque_bombodega.xlsx")
ws=wb.active
for col in ws.columns:
    max_lerngth = 0
    column = col[0].column_letter
    for cell in col :
        try:
            if len(str(cell.value)) > max_lerngth:
                max_lerngth = len(cell.value)
        except:
            pass
    adjusted_width = (max_lerngth + 2)*1.2
    ws.column_dimensions[column].width = adjusted_width
wb.save('controle_de_estoque_bombodega.xlsx')'''


import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter 
from datetime import datetime,timedelta

#A cada 30 dias atualiza a coluna
def adicionar_coluna_periodica(arquivo):
    nome_coluna= datetime.now().strftime('%d / %m / %Y')
    # Carrega o arquivo existente
    wb = load_workbook(arquivo)

    # Seleciona a planilha ativa
    ws = wb.active

    # Obtém a data atual
    data_atual = datetime.now()

    # Verifica se já passaram 30 dias desde a última adição da coluna
    if 'data_ultima_coluna' not in ws.sheet_properties:
        ws.sheet_properties['data_ultima_coluna'] = data_atual - timedelta(days=30)
    
    data_ultima_coluna = ws.sheet_properties['data_ultima_coluna']

    if (data_atual - data_ultima_coluna).days >= 30:
        # Atualiza a data da última coluna adicionada
        ws.sheet_properties['data_ultima_coluna'] = data_atual

        # Obtém o número da última coluna
        num_coluna = ws.max_column

        # Cria uma nova coluna com o nome especificado
        nova_coluna = num_coluna + 1
        letra_coluna = get_column_letter(nova_coluna)
        ws[f'{letra_coluna}1'] = nome_coluna

        # Obtém o valor da coluna 'total_de_produtos'
        valor_total = 0
        for cell in ws['K'][1:]:
            valor_total += cell.value

        # Insere o valor da coluna 'total_de_produtos' na nova coluna
        ws[f'{letra_coluna}2'] = valor_total

        # Salva o arquivo
        wb.save(arquivo)
        print(f"Nova coluna '{nome_coluna}' adicionada com sucesso.")
    else:
        print("Ainda não se passaram 30 dias desde a última adição de coluna.")







# Mostrar conteudo
def mostrar(txt):
    wb = load_workbook(txt)
    ws= wb.active
    for row in ws.iter_rows(values_only=True):
        print(row)

#Gerador ID
def gerador_id (txt , categoria):
    from openpyxl import load_workbook
    wb = load_workbook(txt)
    ws = wb.active
    #ver se o tipo ja existe
    for row in ws.iter_rows(min_row = 2 , max_col = 9 , max_row = ws.max_row , values_only = True):
        if row[1] == categoria :
            return row[0]
    novo_id = 1
    for row in ws.iter_rows(min_row = 2 , max_col = 1 , max_row = ws.max_row , values_only = True):
        if row[0] == novo_id :
            novo_id = row[0]+1
    return novo_id



def cadastrar_produto():
    from openpyxl.styles import PatternFill
    produto = entry_nome_produto.get()
    quantidade = int(entry_quantidade.get())
    preco_bruto = float(entry_preco_bruto.get())
    preco_final = float(entry_preco_final.get())
    fornecedor = entry_fornecedor.get()
    data_entrada = datetime.now().strftime('%d / %m / %Y  %H : %M')
    data_validade = entry_data_validade.get()
    categoria = entry_categoria.get()


    # Carrega o arquivo existente
    wb = load_workbook('controle_de_estoque_bombodega.xlsx')

    # Seleciona a planilha ativa
    ws = wb.active
    # adiciona novo produto
    while True:
        ok = False
        novo = ( gerador_id('controle_de_estoque_bombodega.xlsx',categoria),f'{produto}',quantidade,
            preco_bruto,preco_final,f'{fornecedor}',data_entrada,data_validade,f'{categoria}'
           )
        ws.append(novo)
        wb.save('controle_de_estoque_bombodega.xlsx')
        ok = True
    #coloca produtos no historico
        if ok == True :
            wb.active = wb['Historico']
            wsd = wb.active
            new =(f'{produto}',quantidade,data_entrada,f'{categoria}','entrada')
            wsd.append(new)
            wb.save('controle_de_estoque_bombodega.xlsx')
        break
    #ajusta as colunas
    for col in ws.columns and wsd.columns:
        max_lerngth = 0
        column = col[0].column_letter
        for cell in col :
            try:
                if len(str(cell.value)) > max_lerngth:
                    max_lerngth = len(cell.value)
            except:
                pass
        adjusted_width = (max_lerngth + 2)*1.2
        ws.column_dimensions[column].width = adjusted_width
    wb.save('controle_de_estoque_bombodega.xlsx')
    messagebox.showinfo('Sucesso', f'Produto "{produto}" cadastrado com sucesso!')

def excluir_item():
    from openpyxl.styles import PatternFill
    #carrega tabela
    wb = load_workbook ('controle_de_estoque_bombodega.xlsx')
    ws = wb['Sheet']
    #pega a data atua
    data_atual = datetime.now().strftime('%d / %m / %Y  %H %M')
    #vai iterar e separar a planilha em listas
    intervalo = ws.iter_rows(max_row= ws.max_row , max_col=3)
    intervalo = [[x.value for x in row]for row in intervalo]
    #vai procurar o produto
    if intervalo [0][1] == 'produto':
        for l in intervalo:
    #se achar vai calcular e subistituir a quantidade
            if l[1] == entry_nome_produto.get():
                quant = l[2]
                nova_quantia= max(quant-int(entry_quantidade.get()),0) 
                ws.cell(row =l[0],column=3).value= nova_quantia
    wb.save('controle_de_estoque_bombodega.xlsx')
    
     #Adiciona a exclusao no historico
    wsd= wb['Historico']
    wsd.append([entry_nome_produto.get(),entry_quantidade.get(),data_atual,entry_categoria.get(),"saida"])
    #Pinta a linha de vermelha
    for cell in wsd[wsd.max_row]:
        cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000",fill_type = "solid") 
    wb.save('controle_de_estoque_bombodega.xlsx') 
    messagebox.showinfo('Sucesso', f'Produto "{entry_nome_produto.get()}" excluido com sucesso!')

# Cria a janela principal
root = tk.Tk()
root.title('Cadastro de Produto')
root.configure(background='plum2' )
root.geometry("700x500")
root.resizable(True,True)

# Cria os campos de entrada
tk.Label(root, text='Produto:',bd=4,bg='mistyrose',
         highlightbackground="maroon2",highlightthickness= 2).place(relx=0.09, rely=0.04,relwidth= 0.4, relheight=0.05)
entry_nome_produto = tk.Entry(root)
entry_nome_produto.place(relx=0.5, rely=0.04,relwidth=0.4, relheight=0.05)

tk.Label(root, text='Quantidade:',bd=4,bg='mistyrose',
         highlightbackground="maroon2",highlightthickness= 2).place(relx=0.09, rely=0.10,relwidth= 0.4, relheight=0.05)
entry_quantidade = tk.Entry(root)
entry_quantidade.place(relx=0.5, rely=0.1,relwidth=0.4, relheight=0.05)

tk.Label(root, text='Preço bruto (R$):',bd=4,bg='mistyrose',
         highlightbackground="maroon2",highlightthickness= 2).place(relx=0.09, rely=0.16,relwidth= 0.4, relheight=0.05)
entry_preco_bruto = tk.Entry(root)
entry_preco_bruto.place(relx=0.5, rely=0.16,relwidth=0.4, relheight=0.05)

tk.Label(root, text='Preço final (R$):',bd=4,bg='mistyrose',
         highlightbackground="maroon2",highlightthickness= 2).place(relx=0.09, rely=0.22,relwidth= 0.4, relheight=0.05)
entry_preco_final = tk.Entry(root)
entry_preco_final.place(relx=0.5, rely=0.22,relwidth=0.4, relheight=0.05)

tk.Label(root, text='fornecedor:',bd=4,bg='mistyrose',
         highlightbackground="maroon2",highlightthickness= 2).place(relx=0.09, rely=0.28,relwidth= 0.4, relheight=0.05)
entry_fornecedor = tk.Entry(root)
entry_fornecedor.place(relx=0.5, rely=0.28,relwidth=0.4, relheight=0.05)

tk.Label(root, text='Data de valdade :',bd=4,bg='mistyrose',
         highlightbackground="maroon2",highlightthickness= 2).place(relx=0.09, rely=0.34,relwidth= 0.4, relheight=0.05)
entry_data_validade = tk.Entry(root)
entry_data_validade.place(relx=0.5, rely=0.34,relwidth=0.4, relheight=0.05)

tk.Label(root, text='Categoria :',bd=4,bg='mistyrose',
         highlightbackground="maroon2",highlightthickness= 2).place(relx=0.09,
                                                             rely=0.4,relwidth= 0.4, relheight=0.05)
entry_categoria = tk.Entry(root)
entry_categoria.place(relx=0.5, rely=0.4,relwidth=0.4, relheight=0.05)

# Botão para cadastrar o produto
btn_cadastrar = tk.Button(root, text='Cadastrar Produto', command=cadastrar_produto,background="gray70",bd = 2)
btn_cadastrar.place(relx=0.25, rely=0.46,relwidth=0.5, relheight=0.05)
#Botao para excluir o produto
btn_excluir = tk.Button(root,text = 'Excluir Produto',command=excluir_item,bg= 'red',bd= 2,fg="white")
btn_excluir.place(relx=0.25, rely=0.525,relwidth=0.5, relheight=0.05)
# preview da tabela
def produtos_vencendo():
    wb = load_workbook('controle_de_estoque_bombodega.xlsx')
    ws = wb.active

    data_atual = datetime.now()
    proximo_mes = data_atual.replace(month=data_atual.month + 1)
    produtos_vencer = []
    for row in ws.iter_rows(min_row = 2 , max_row = ws.max_row , values_only = True):
        dataValidade = row[7]
        if dataValidade.month == proximo_mes.month and dataValidade.year == proximo_mes.year :
            produtos_vencer.append(row[0,1,2,6,7])
    texto_produtos = "\n".join(produtos_vencer)
tabela =tk.Frame(root,bd=4,bg='#dfe3ee',
         highlightbackground="#759fe6",highlightthickness= 4).place(relx=0.09, rely=0.6,relwidth=0.81, relheight=0.35)
lista = ttk.Treeview(tabela,height= 3 , columns = ("ID","NOME","QUANTIDADE","DATA DE VALIDADE"),show = 'headings')
#Cabeçalho lista
lista.heading('ID',text="ID")
lista.heading('NOME',text="NOME")
lista.heading('QUANTIDADE',text="QUANTIDADE")
lista.heading('DATA DE VALIDADE',text="VALIDADE")
#Correção tamanho
lista.column("ID", width= 10)
lista.column("NOME", width= 10,anchor= tk.CENTER )
lista.column("QUANTIDADE", width= 10,anchor= tk.CENTER)
lista.column("DATA DE VALIDADE", width= 10,anchor= tk.CENTER)
#posição da lista
lista.place(relx=0.135, rely=0.64,relwidth=0.68, relheight=0.276)
#Barra rolagem
scroll_lista =tk.Scrollbar(tabela,orient = "vertical")
lista.configure(yscroll= scroll_lista.set)
scroll_lista.place(relx=0.815 , rely = 0.64 ,relwidth= 0.04 , relheight = 0.276)
# Inicia a interface
root.mainloop()

